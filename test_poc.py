import pprint

import openpyxl
from binance.client import Client

import config
import helper

client = Client(config.API_KEY, config.API_SECRET)

exchange_info = client.get_exchange_info()

# Loading the workbook for SST
wb = openpyxl.load_workbook('sst_1.xlsx')

# Loading the Main sheet from Workbook
sh1 = wb['Main']

# storing the total number of row count in row variable
row = sh1.max_row

# storing the total number of cloumn count in col variable
col = sh1.max_column

# Loading the workbook for SST
wb = openpyxl.load_workbook('sst_1.xlsx')
helper.fetch_all_symbols_from_main()
helper.fetch_all_sym_where_we_have_positions(helper.get_exchange())

# ccxt_exchange=helper.get_exchange()
# since = ccxt_exchange.parse8601('2016-12-12T00:00:00')
# trades = ccxt_exchange.fetch_my_trades('SHIB/USDT',since)
# print('Fetched', len(trades), 'trades')
# if len(trades):
#     last_trade = trades[len(trades) - 2]
#     from_id = last_trade['id']
#     pprint.pprint(last_trade)

#pprint.pprint(trades)