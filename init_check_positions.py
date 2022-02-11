import pprint

import openpyxl
from binance import Client
from config import *


import config
import helper

def init():
    # opening client connection to binance
    client = Client(config.API_KEY, config.API_SECRET)

    # Loading the workbook for SST
    wb = openpyxl.load_workbook(WORKBOOK_FOR_SST)
    sh3 = wb[WORKBOOK_INVESTMENT_SHEET]
    sh2 = wb[WORKBOOK_TRACKING_SHEET]
    helper.fetch_all_symbols_from_main()
    helper.fetch_all_sym_where_we_have_positions(helper.get_exchange())
    ccxt_exchange = helper.get_exchange()
    since = ccxt_exchange.parse8601('2020-12-12T00:00:00')
    for i in range(1, len(helper.sym_where_we_have_position)):
        sh3.cell(i + 1, 1).value = list(helper.sym_where_we_have_position)[i]
        spot = list(filter(lambda x: x.get('symbol') == list(helper.sym_where_we_have_position)[i], client.get_all_tickers()))
        spot_price = spot[0].get('price')
        sh3.cell(i + 1, 2).value = spot_price

        sym=list(helper.sym_where_we_have_position)[i]
        sym_required = str(sym).split('USDT')[0]+'/USDT'
        trades = ccxt_exchange.fetch_my_trades(sym_required, since)
        print('Fetched', len(trades), 'trades')
        last_buy_trade = {}
        if len(trades):
            for j in range(0,len(trades)) :
                last_trade = trades[len(trades) - j-1]
                if last_trade['side']=='buy':
                    last_buy_trade = last_trade
                    break
            last_trade = trades[len(trades) -j -1]
            #pprint.pprint(last_trade)
        sh3.cell(i + 1, 3).value = last_buy_trade['price']
        sh3.cell(i + 1, 4).value = last_buy_trade['datetime']
        sh3.cell(i+1,5).value = ((float(spot_price) - float(last_buy_trade['price']))/float(last_buy_trade['price']))*100


    #Loop to delete rows in Tracking sheet
    row2 = sh2.max_row
    for i in range(1, row2):
        sym = sh2.cell(i + 1, 1).value
        if sym in helper.sym_where_we_have_position:
            sh2.delete_rows(idx = i+1)
            print(f'Position in {sym} therefore deleting it from Tracking sheet')


    wb.save(WORKBOOK_FOR_SST)
def main():
    init()

main()