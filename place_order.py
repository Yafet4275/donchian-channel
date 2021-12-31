import openpyxl
from binance import Client
from binance.enums import *

import config
import helper


def init():
    #opening client connection to binance
    client = Client(config.API_KEY, config.API_SECRET)
    # info = client.get_open_orders(symbol = 'LTCUSDT')
    #
    # for inf in info:
    #     print(inf)
    # print(info)


    # Loading the workbook for SST
    wb = openpyxl.load_workbook('sst.xlsx')

    # Loading the Tracking sheet from Workbook
    sh2 = wb['Tracking']

    #storing the total number of row count in row variable
    row = sh2.max_row

    # Iterating all the rows in sst.xlsx
    for i in range(1, row):
        # Skipping the header using + 1 and reading the first column of each row in variable sym
        sym = sh2.cell(i + 1, 1).value
        away_per = sh2.cell(i + 1, 4).value

        #First condition for Buy is Away % should be between 0 to 5
        if 0<= away_per <= 15:
            #If the GTT Order Price column in excel is None then the 20 Day High is the GTT Price
            if sh2.cell(i + 1, 6).value is None:
                gtt_order_price = sh2.cell(i + 1, 3).value
                # Updating the 20 DH in the GTT order price column
                sh2.cell(i + 1, 6).value = gtt_order_price

                # Placing a buy order and this is a fresh scenario
                helper.buy(client, sym, gtt_order_price)
                print(f'Buy order placed for {sym}')
            else:
                if sh2.cell(i + 1, 8).value == 'Y':
                    print(f'Updating the previous GTT order for {sym}')
                    gtt_order_price = sh2.cell(i + 1, 3).value
                    sh2.cell(i + 1, 6).value = gtt_order_price
                    sh2.cell(i + 1, 8).value =''
                    helper.buy(client, sym, gtt_order_price)
                else:
                    print(f'An order already exists for {sym}')

    #Saving the workbook
    wb.save('sst.xlsx')

def main():
    init()

main()